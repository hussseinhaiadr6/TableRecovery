import os

def remove_empty_directories(directory):
    for root, dirs, files in os.walk(directory):
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            if not os.listdir(dir_path):
                os.rmdir(dir_path)
                print(f'Removed empty directory: {dir_path}')


import os
import shutil
import re


def flatten_and_rename(root_directory):
    for chunk_dir in os.listdir(root_directory):
        chunk_path = os.path.join(root_directory, chunk_dir)

        if os.path.isdir(chunk_path):
            image_counter = {}
            print(f"Processing directory: {chunk_path}")

            for root, dirs, files in os.walk(chunk_path, topdown=False):
                for file in files:
                    if file.endswith(('.jpg', '.jpeg', '.png')):
                        page_dir = os.path.basename(root)
                        match = re.match(r'page(\d+)', page_dir)
                        if match:
                            page_number = match.group(1)
                            table_number = image_counter.get(page_number, 0)
                            new_name = f'{chunk_dir}_Page{page_number}_Table{table_number}.jpg'
                            old_path = os.path.join(root, file)
                            new_path = os.path.join(chunk_path, new_name)
                            shutil.move(old_path, new_path)
                            print(f'Renamed {old_path} to {new_path}')
                            image_counter[page_number] = table_number + 1

                # Remove empty directories
                for dir_name in dirs:
                    dir_path = os.path.join(root, dir_name)
                    if not os.listdir(dir_path):
                        os.rmdir(dir_path)
                        print(f'Removed empty directory: {dir_path}')

from openpyxl import load_workbook,Workbook
import openpyxl
import os
def combine_xlsx_files(input_folder, output_filename):

  # Create a new workbook for the combined data
  wb_out = Workbook()
  ws_out = wb_out.active
  last_written_row=1
  Xlsx_files = deep_search_file_type(input_folder, ".xlsx")
  # Loop through XLSX files in the folder
  i=0
  for Xlsx_file in Xlsx_files  :

      print(Xlsx_file)
      # print(filename)
      # Load the current XLSX file
      wb_in = load_workbook(filename=Xlsx_file)
      ws_in = wb_in.active

      # Get last row number of the current file
      last_row_in = ws_in.max_row
      # print(last_row_in)

      # Copy data from the current file to the output sheet
      for row in ws_in.iter_rows(min_row=1, max_row=last_row_in, values_only=True):
        # Get the next available row in the output sheet
        next_row_out = 1
        print(row)

        # Copy the data row and insert four empty rows below
        for col, value in enumerate(row, 1):
          # if value== 'GTI322 - GLEM':
          #   print('break here')
          ws_out.cell(row=last_written_row, column=col).value = value
          next_row_out+=1
          print(value,col)
        last_written_row =last_written_row+1
         # next_row_out += 4  # Move down by four rows for empty space
      ws_out.insert_rows(last_written_row + 1, 4)  # Insert 4 empty rows
      # print('break here')
      last_written_row =last_written_row+ 4
      i=1

  # Save the combined workbook
  wb_out.save("combined_data.xlsx")

  print(f"Combined XLSX files into '{output_filename}'.")

# Replace with the actual paths


def is_xlsx_empty(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active  # Get the active sheet

        # Check for empty cells in the first sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    return False  # Non-empty cell found, so not empty

        # If no non-empty cells found, check for empty rows
        if all(row[0].value is None for row in sheet.iter_rows()):
            return True  # All rows empty, so empty sheet
        else:
            return False  # Sheet contains empty rows but also non-empty cells
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return False  # Consider this empty for consistency

def is_xlsx_empty_wrap(filename):
  if is_xlsx_empty(filename):
      print('removed file:',filename)
      os.remove(filename)


from PyPDF2 import PdfReader, PdfWriter

def split_pdf(input_filename, output_dir, pages_per_chunk):
  # Ensure output directory exists
  os.makedirs(output_dir, exist_ok=True)  # Create directory if it doesn't exist

  with open(input_filename, 'rb') as input_file:
    pdf_reader = PdfReader(input_file)
    num_pages = len(pdf_reader.pages)


    # Calculate the total number of chunks
    num_chunks = num_pages // pages_per_chunk + (num_pages % pages_per_chunk > 0)

    for chunk_number in range(num_chunks):
      start_page = chunk_number * pages_per_chunk
      end_page = min(start_page + pages_per_chunk, num_pages)  # Handle last chunk

      pdf_writer = PdfWriter()
      for page_num in range(start_page, end_page):
        pdf_writer.add_page(pdf_reader.pages[page_num])

      output_filename = os.path.join(output_dir, f"chunk_{chunk_number+1}.pdf")  # 1-based indexing
      if os.path.exists(output_filename):
        # Delete the existing file
        os.remove(output_filename)
        print(f"Removing {output_filename}")
      with open(output_filename, 'wb') as output_file:
        pdf_writer.write(output_file)

  print(f"PDF split into {num_chunks} chunks in directory: {output_dir}")




from pdf2image import convert_from_path
import os
poppler_path = "./poppler/poppler-24.02.0/Library/bin"
def pdftoImages(Source_dir, Save_folder=""):
    os.makedirs("./images", exist_ok=True)  # Create directory if it doesn't exist
    try:
        for root, dirs, files in os.walk(Source_dir):
            for filename in files:
                if filename.endswith(".pdf"):
                    os.makedirs(Save_folder + "/" + filename.split('.')[0],exist_ok=True)
                    print(filename)
                    images = convert_from_path(root+"/"+filename, poppler_path=poppler_path)
                    for i in range(len(images)):

                        images[i].save(Save_folder+'/'+filename.split('.')[0]+'/page' + str(i + 1) + '.jpg', 'JPEG')

    except FileNotFoundError:
        print("ERROR: Create a folder Called Images to continue in Task 1 Directory")



def get_all_file_paths(directory):
  file_paths = []
  for filename in os.listdir(directory):
    # Construct the full path by joining the directory path and filename
    full_path = os.path.join(directory, filename)

    # Check if it's a file (not a directory) using os.path.isfile
    if os.path.isfile(full_path):
      file_paths.append(full_path)

  return file_paths




def get_all_dirs_paths(directory):
  dir_paths = []
  for filename in os.listdir(directory):
    # Construct the full path by joining the directory path and filename
    full_path = os.path.join(directory, filename)

    # Check if it's a file (not a directory) using os.path.isfile
    if os.path.isdir(full_path):
      dir_paths.append(full_path)

  return dir_paths


def deep_search_file_type(Source_dir, extension):
  list_of_files=[]
  for root, dirs, files in os.walk(Source_dir):
    for filename in files:
      if filename.endswith(extension):
        list_of_files.append(root+"/"+filename)
  return list_of_files
# Replace this with your actual path
"""tables_detected_path = "./Tables_Detected"
flatten_and_rename(tables_detected_path)
"""