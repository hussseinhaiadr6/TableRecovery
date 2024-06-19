from openpyxl import load_workbook,Workbook
import openpyxl
import os
from variables import get_path
from getAllFileInPath import deep_search_file_type, get_all_file_paths, get_all_dirs_paths

path=get_path()
def combine_xlsx_files(input_folder, output_filename):

  # Create a new workbook for the combined data
  wb_out = Workbook()
  ws_out = wb_out.active
  last_written_row=1
  Xlsx_files = deep_search_file_type(input_folder, ".xlsx")
  # Loop through XLSX files in the folder
  i=0
  for Xlsx_file in Xlsx_files  :

      print(Xlsx_file)
      # print(filename)
      # Load the current XLSX file
      wb_in = load_workbook(filename=Xlsx_file)
      ws_in = wb_in.active

      # Get last row number of the current file
      last_row_in = ws_in.max_row
      # print(last_row_in)

      # Copy data from the current file to the output sheet
      for row in ws_in.iter_rows(min_row=1, max_row=last_row_in, values_only=True):
        # Get the next available row in the output sheet
        next_row_out = 1
        print(row)

        # Copy the data row and insert four empty rows below
        for col, value in enumerate(row, 1):
          # if value== 'GTI322 - GLEM':
          #   print('break here')
          ws_out.cell(row=last_written_row, column=col).value = value
          next_row_out+=1
          print(value,col)
        last_written_row =last_written_row+1
         # next_row_out += 4  # Move down by four rows for empty space
      ws_out.insert_rows(last_written_row + 1, 4)  # Insert 4 empty rows
      # print('break here')
      last_written_row =last_written_row+ 4
      i=1

  # Save the combined workbook
  wb_out.save("combined_data.xlsx")

  print(f"Combined XLSX files into '{output_filename}'.")

# Replace with the actual paths