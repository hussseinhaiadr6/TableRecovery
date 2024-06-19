import openpyxl
import os
def is_xlsx_empty(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active  # Get the active sheet

        # Check for empty cells in the first sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    return False  # Non-empty cell found, so not empty

        # If no non-empty cells found, check for empty rows
        if all(row[0].value is None for row in sheet.iter_rows()):
            return True  # All rows empty, so empty sheet
        else:
            return False  # Sheet contains empty rows but also non-empty cells
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return False  # Consider this empty for consistency

def is_xlsx_empty_wrap(filename):
  if is_xlsx_empty(filename):
      print('removed file:',filename)
      os.remove(filename)
